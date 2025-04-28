from flask import Blueprint, request, jsonify, send_file
from extensions import db
from models.recipe import Recipe, RecipeMaterial, RecipeSchema
from models.production import ProductionOrder
from models.user import User
from sqlalchemy.exc import IntegrityError
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from barcode import Code128
from barcode.writer import ImageWriter
from PIL import Image as PILImage
import io, os, tempfile
from werkzeug.exceptions import BadRequest
import logging

recipe_bp = Blueprint("recipe", __name__)
logging.basicConfig(level=logging.DEBUG)

@recipe_bp.route("/recipes/export/barcodes", methods=["GET"])
def export_recipes_excel_with_barcodes():
    try:
        recipes = Recipe.query.all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Recipes with Barcodes"
        ws.append(["Name", "Code", "Barcode ID", "Scannable Barcode"])

        row_number = 2
        for recipe in recipes:
            if recipe.barcode_id:
                barcode_id = recipe.barcode_id
                try:
                    temp_dir = tempfile.gettempdir()
                    filename = f"{barcode_id}"
                    filepath = os.path.join(temp_dir, f"{filename}.png")
                    code128 = Code128(barcode_id, writer=ImageWriter())
                    code128.save(os.path.join(temp_dir, filename))
                    image = PILImage.open(filepath)
                    image = image.resize((200, 60))
                    image.save(filepath)

                    ws.cell(row=row_number, column=1, value=recipe.name)
                    ws.cell(row=row_number, column=2, value=recipe.code)
                    ws.cell(row=row_number, column=3, value=barcode_id)

                    img = ExcelImage(filepath)
                    img.width = 150
                    img.height = 50
                    ws.add_image(img, f"D{row_number}")

                    os.remove(filepath)
                    row_number += 1

                except Exception as e:
                    logging.error(f"Failed to generate barcode for {barcode_id}: {e}")
                    ws.cell(row=row_number, column=1, value=recipe.name)
                    ws.cell(row=row_number, column=2, value=recipe.code)
                    ws.cell(row=row_number, column=3, value=barcode_id)
                    row_number += 1

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        return send_file(
            stream,
            download_name="recipes_with_barcodes.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@recipe_bp.route("/recipes", methods=["GET"])
def get_recipes():
    recipes = Recipe.query.all()
    result = [
        {
            "recipe_id": recipe.recipe_id,
            "name": recipe.name,
            "code": recipe.code,
            "description": recipe.description,
            "version": recipe.version,
            "status": recipe.status,
            "created_by": recipe.created_by,
            "created_at": recipe.created_at,
            "no_of_materials": recipe.no_of_materials,
        }
        for recipe in recipes
    ]
    return jsonify(result)

@recipe_bp.route("/recipes/paginated", methods=["GET"])
def get_paginated_recipes():
    page = request.args.get("page", 1, type=int)
    page_size = request.args.get("page_size", 10, type=int)
    pagination = Recipe.query.paginate(page=page, per_page=page_size, error_out=False)
    result = [
        {
            "recipe_id": recipe.recipe_id,
            "name": recipe.name,
            "code": recipe.code,
            "description": recipe.description,
            "version": recipe.version,
            "status": recipe.status,
            "created_by": recipe.created_by,
            "created_at": recipe.created_at,
            "no_of_materials": recipe.no_of_materials,
        }
        for recipe in pagination.items
    ]
    return jsonify({
        "recipes": result,
        "total": pagination.total,
        "page": pagination.page,
        "pages": pagination.pages,
        "per_page": pagination.per_page,
        "has_next": pagination.has_next,
        "has_prev": pagination.has_prev
    })

@recipe_bp.route("/recipes/all", methods=["GET"])
def get_all_recipes():
    recipes = Recipe.query.all()
    recipe_schema = RecipeSchema(many=True)
    return jsonify(recipe_schema.dump(recipes)), 200



@recipe_bp.route("/recipes", methods=["POST"])
def create_recipe():
    data = request.get_json()

    required_fields = ["name", "code", "version", "created_by"]
    for field in required_fields:
        if not data.get(field):
            return jsonify({"error": f"'{field}' is required."}), 400

    # ✅ Validate status
    status = data.get("status", "Unreleased")
    valid_statuses = ["Released", "Unreleased"]
    if status not in valid_statuses:
        return jsonify({"error": f"Invalid status value: {status}"}), 400

    # Validate no_of_materials
    no_of_materials = data.get("no_of_materials")
    if no_of_materials is not None:
        try:
            no_of_materials = int(no_of_materials)
            if no_of_materials < 0:
                raise ValueError
        except (ValueError, TypeError):
            return jsonify({"error": "no_of_materials must be a non-negative integer."}), 400


    # ✅ Validate user existence
    user = db.session.get(User, data["created_by"])
    if not user:
        return jsonify({"error": "User not found."}), 400

    new_recipe = Recipe(
        name=data["name"],
        code=data["code"],
        description=data.get("description"),
        version=data["version"],
        status=status,
        created_by=data["created_by"],
        barcode_id=data.get("barcode_id"),
        no_of_materials=no_of_materials
    )

    db.session.add(new_recipe)

    try:
        db.session.commit()
    except IntegrityError as e:
        db.session.rollback()
        if "Duplicate entry" in str(e.orig):
            return jsonify({"error": "Duplicate entry: code or barcode_id already exists."}), 400
        return jsonify({"error": "Database error occurred."}), 500
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

    return jsonify({"message": "Recipe created successfully!"}), 201

@recipe_bp.route("/recipes/<int:recipe_id>", methods=["GET"])
def get_recipe(recipe_id):
    recipe = Recipe.query.get(recipe_id)
    if not recipe:
        return jsonify({"error": "Recipe not found"}), 404

    result = {
        "recipe_id": recipe.recipe_id,
        "name": recipe.name,
        "code": recipe.code,
        "description": recipe.description,
        "version": recipe.version,
        "status": recipe.status,
        "created_by": recipe.created_by,
        "no_of_materials" : recipe.no_of_materials
    }
    return jsonify(result)



@recipe_bp.route("/recipes/<int:recipe_id>", methods=["PUT"])
def update_recipe(recipe_id):
    try:
        recipe = Recipe.query.get(recipe_id)
        if not recipe:
            return jsonify({"message": "Recipe not found"}), 404

        data = request.get_json()

        # Update fields if provided, otherwise leave unchanged
        recipe.name = data.get("name", recipe.name)
        recipe.code = data.get("code", recipe.code)
        recipe.description = data.get("description", recipe.description)
        recipe.version = data.get("version", recipe.version)
        recipe.status = data.get("status", recipe.status)
        recipe.no_of_materials = data.get("no_of_materials", recipe.no_of_materials)

        # 💡 Added sequence field update
        if "sequence" in data:
            recipe.sequence = data["sequence"]

        db.session.commit()
        return jsonify({"message": "Recipe updated successfully"}), 200

    except Exception as e:
        db.session.rollback()  # Rollback in case of any error
        app.logger.error(f"Error updating recipe {recipe_id}: {str(e)}")
        return jsonify({"message": "An error occurred while updating the recipe."}), 500


@recipe_bp.route("/recipes/<int:recipe_id>", methods=["DELETE"])
def delete_recipe(recipe_id):
    try:
        # Step 1: Delete related records in `production_order`
        db.session.query(ProductionOrder).filter(ProductionOrder.recipe_id == recipe_id).delete(synchronize_session=False)

        # Step 2: Delete related `recipe_material` records before setting NULL
        db.session.query(RecipeMaterial).filter(RecipeMaterial.recipe_id == recipe_id).delete(synchronize_session=False)

        # Step 3: Now delete the recipe itself
        recipe = Recipe.query.get(recipe_id)
        if not recipe:
            return jsonify({"message": "Recipe not found"}), 404

        db.session.delete(recipe)
        db.session.commit()  # Commit the deletions

        return jsonify({"message": "Recipe and related data deleted successfully"}), 200

    except IntegrityError as e:
        db.session.rollback()  # Rollback the session if an integrity error occurs
        return jsonify({"message": "Integrity error, check related records for consistency"}), 400
    except Exception as e:
        db.session.rollback()  # Rollback the session in case of any other errors
        return jsonify({"message": "An error occurred while deleting the recipe"}), 500

### 🚀 RECIPE MATERIAL ROUTES ###
@recipe_bp.route("/recipe_materials", methods=["POST"])
def create_or_update_recipe_material():
    try:
        data = request.get_json()

        if not data:
            raise BadRequest("No input data provided.")

        recipe_id = data.get("recipe_id")
        material_id = data.get("material_id")
        set_point = data.get("set_point")
        status = data.get("status")
        bucket_id = data.get("bucket_id")  # ✅ Extract bucket_id from request
        
        # Use scale flag
        use_scale = data.get("use_scale", False)

        # Get actual weight
        if use_scale:
            from models.scale import ScaleClient
            scale_client = ScaleClient()
            actual = scale_client.get_net_weight()
            if actual is None:
                return jsonify({
                    "error": "Failed to read weight from scale. Please check scale connection."
                }), 500
        else:
            actual = data.get("actual")

        if not recipe_id or not material_id or set_point is None or not status:
            raise BadRequest("Missing required fields. Please provide recipe_id, material_id, set_point, and status.")
        if actual is None:
            raise BadRequest("Actual weight is required. Either provide 'actual' value or set 'use_scale' to true.")
        if not isinstance(recipe_id, int) or not isinstance(material_id, int):
            raise BadRequest("recipe_id and material_id must be integers.")
        if not isinstance(set_point, (int, float)) or not isinstance(actual, (int, float)):
            raise BadRequest("set_point and actual must be numeric values.")

        # ✅ Optional: validate bucket_id exists
        if bucket_id is not None:
            from models import StorageBucket
            if not StorageBucket.query.get(bucket_id):
                raise BadRequest("Invalid bucket_id. Bucket not found.")

        # Calculate margin
        margin = 0.0 if set_point == 0 else round(((float(set_point) - float(actual)) / float(set_point)) * 100, 2)

        # Check for existing recipe material
        existing_recipe_material = RecipeMaterial.query.filter_by(recipe_id=recipe_id, material_id=material_id).first()

        if existing_recipe_material:
            # Update
            existing_recipe_material.set_point = set_point
            existing_recipe_material.actual = actual
            existing_recipe_material.margin = margin
            existing_recipe_material.status = status
            existing_recipe_material.bucket_id = bucket_id  # ✅ Save bucket_id
            db.session.commit()

            return jsonify({
                "message": "Recipe material updated successfully!",
                "margin": f"{margin}%",
                "actual": actual,
                "scale_used": use_scale
            }), 200
        else:
            # Create
            new_recipe_material = RecipeMaterial(
                recipe_id=recipe_id,
                material_id=material_id,
                set_point=set_point,
                actual=actual,
                margin=margin,
                status=status,
                bucket_id=bucket_id  # ✅ Save bucket_id
            )
            db.session.add(new_recipe_material)
            db.session.commit()

            return jsonify({
                "message": "Recipe material created successfully!",
                "margin": f"{margin}%",
                "actual": actual,
                "scale_used": use_scale
            }), 201

    except BadRequest as e:
        logging.error(f"Bad request: {e.description}")
        return jsonify({"error": e.description}), 400
    except Exception as e:
        logging.error(f"Unexpected error: {str(e)}")
        return jsonify({"error": "An unexpected error occurred. Please try again later."}), 500



@recipe_bp.route("/recipe_materials", methods=["GET"])
def get_recipe_materials():
    materials = RecipeMaterial.query.all()
    result = [
        {
            "recipe_material_id": mat.recipe_material_id,
            "recipe_id": mat.recipe_id,
            "material_id": mat.material_id,
            "set_point": str(mat.set_point) if mat.set_point is not None else None,
            "actual": str(mat.actual) if mat.actual is not None else None,
            "margin": str(mat.margin) if mat.margin is not None else None  # Include margin field
        }
        for mat in materials
    ]
    return jsonify(result)

@recipe_bp.route("/recipe_materials/<int:recipe_material_id>", methods=["PUT"])
def update_recipe_material(recipe_material_id):
    material = RecipeMaterial.query.get(recipe_material_id)
    if not material:
        return jsonify({"message": "Recipe material not found"}), 404

    data = request.get_json()
    material.material_id = data.get("material_id", material.material_id)
    material.set_point = data.get("set_point", material.set_point)
    # material.actual = data.get("actual", material.actual)

    db.session.commit()
    return jsonify({"message": "Recipe material updated successfully"}), 200

@recipe_bp.route("/recipe_materials/<int:recipe_id>", methods=["GET"])
def get_recipe_materials_by_recipe_id(recipe_id):
    materials = RecipeMaterial.query.filter_by(recipe_id=recipe_id).all()
    if not materials:
        return jsonify({"message": "No materials found for this recipe"}), 404
    
    result = [
        {
            "recipe_material_id": mat.recipe_material_id,
            "recipe_id": mat.recipe_id,
            "material_id": mat.material_id,
            "set_point": str(mat.set_point) if mat.set_point is not None else None,
            "actual": str(mat.actual) if mat.actual is not None else None,
            "margin": str(mat.margin) if mat.margin is not None else None
        }
        for mat in materials
    ]
    return jsonify(result)


@recipe_bp.route("/recipe_materials/<int:recipe_material_id>", methods=["DELETE"])
def delete_recipe_material(recipe_material_id):
    material = RecipeMaterial.query.get(recipe_material_id)
    if not material:
        return jsonify({"message": "Recipe material not found"}), 404

    db.session.query(RecipeMaterial).filter(RecipeMaterial.recipe_id == recipe_id).update({
            RecipeMaterial.recipe_id: new_recipe_id
        }, synchronize_session=False)
        
    db.session.commit()

    db.session.delete(material)
    db.session.commit()
    return jsonify({"message": "Recipe material deleted successfully"}), 200

@recipe_bp.route("/recipe_materials/<int:recipe_id>/<int:material_id>/capture_weight", methods=["POST"])
def capture_weight_for_recipe_material(recipe_id, material_id):
    try:
        # Get the recipe material
        recipe_material = RecipeMaterial.query.filter_by(
            recipe_id=recipe_id, 
            material_id=material_id
        ).first()
        
        if not recipe_material:
            return jsonify({
                "error": "Recipe material not found"
            }), 404
            
        # Get weight from scale
        from models.scale import ScaleClient
        scale_client = ScaleClient()
        actual = scale_client.get_net_weight()
        
        if actual is None:
            return jsonify({
                "error": "Failed to read weight from scale. Please check scale connection."
            }), 500
            
        # Calculate margin
        set_point = float(recipe_material.set_point) if recipe_material.set_point else 0
        if set_point == 0:
            margin = 0.0
        else:
            margin = round(((set_point - float(actual)) / set_point) * 100, 2)
        
        # Update the recipe material
        recipe_material.actual = actual
        recipe_material.margin = margin
        
        # Update status based on actual weight if needed
        # For example, mark as "created" if weight has been captured
        if recipe_material.status == "pending" or recipe_material.status == "in progress":
            recipe_material.status = "created"
            
        db.session.commit()
        
        return jsonify({
            "message": "Weight captured successfully!",
            "recipe_material_id": recipe_material.recipe_material_id,
            "actual": actual,
            "margin": f"{margin}%",
            "status": recipe_material.status
        }), 200
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error capturing weight: {str(e)}")
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500