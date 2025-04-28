from flask import Flask, Blueprint  , request,current_app, jsonify,abort # type: ignore
from extensions import db
from models.material import Material, MaterialTransaction, MaterialSchema, MaterialTransactionSchema
from models.recipe import RecipeMaterial , Recipe
from sqlalchemy.exc import IntegrityError # type: ignore
from flask import send_file
import io
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
import barcode
from barcode.writer import ImageWriter
from PIL import Image as PILImage
import re
import tempfile
from sqlalchemy.exc import SQLAlchemyError
import logging
from sqlalchemy.orm import aliased




material_bp = Blueprint("materials", __name__)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s]: %(message)s')

material_schema = MaterialSchema()
materials_schema = MaterialSchema(many=True)

transaction_schema = MaterialTransactionSchema()
transactions_schema = MaterialTransactionSchema(many=True)

@material_bp.route("/materials/export/barcodes", methods=["GET"])
def export_materials_excel_with_barcodes():
    try:
        materials = Material.query.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Material Barcodes"
        ws.append(["Title", "Barcode ID", "Scannable Barcode"])

        row_number = 2

        for material in materials:
            if material.barcode_id:
                barcode_id = material.barcode_id

                try:
                    # Create a temp directory and save barcode
                    temp_dir = tempfile.gettempdir()
                    barcode_filename = f"{barcode_id}"
                    barcode_path = os.path.join(temp_dir, f"{barcode_filename}.png")

                    code128 = Code128(barcode_id, writer=ImageWriter())
                    code128.save(os.path.join(temp_dir, barcode_filename))  # Don't include .png here

                    # Resize for Excel
                    image = PILImage.open(barcode_path)
                    image = image.resize((200, 60))
                    image.save(barcode_path)

                    # Fill Excel
                    ws.cell(row=row_number, column=1, value=material.title)
                    ws.cell(row=row_number, column=2, value=barcode_id)

                    img = ExcelImage(barcode_path)
                    img.width = 150
                    img.height = 50
                    ws.add_image(img, f"C{row_number}")

                    # Optional: delete temp image after
                    os.remove(barcode_path)

                    row_number += 1

                except Exception as e:
                    print(f"Failed to generate barcode for {barcode_id}: {e}")
                    # Still add text if image fails
                    ws.cell(row=row_number, column=1, value=material.title)
                    ws.cell(row=row_number, column=2, value=barcode_id)
                    row_number += 1

        # Save to bytes
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        return send_file(
            stream,
            download_name="materials_with_barcodes.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ➤ Create a new Material
@material_bp.route("/materials", methods=["POST"])
def add_material():
    try:
        data = request.get_json()

        # Extract and validate status
        status_value = data.get("status", "Unreleased")
        if status_value not in ["Released", "Unreleased"]:
            return jsonify({"error": "Invalid status value. Allowed values: 'Released', 'Unreleased'"}), 400

        # Convert quantities to float
        current_quantity = float(data.get("current_quantity", 0))
        minimum_quantity = float(data.get("minimum_quantity", 0))
        maximum_quantity = float(data.get("maximum_quantity", 0))

        # Calculate margin safely
        margin = 0.0
        if maximum_quantity > 0:
            margin = round(((maximum_quantity - current_quantity) / maximum_quantity) * 100, 2)

        # Create new Material object
        new_material = Material(
            title=data.get("title"),
            description=data.get("description"),
            unit_of_measure=data.get("unit_of_measure"),
            current_quantity=current_quantity,
            minimum_quantity=minimum_quantity,
            maximum_quantity=maximum_quantity,
            plant_area_location=data.get("plant_area_location"),
            barcode_id=data.get("barcode_id"),
            status=status_value,
            supplier=data.get("supplier"),
            supplier_contact_info=data.get("supplier_contact_info"),
            notes=data.get("notes"),
            margin=margin
        )

        db.session.add(new_material)
        db.session.commit()

        return jsonify({
            "message": "Material added successfully",
            "material": {
                "id": new_material.material_id,
                "title": new_material.title,
                "supplier": new_material.supplier,
                "supplier_contact_info": new_material.supplier_contact_info,
                "notes": new_material.notes,
                "status": new_material.status,
                "margin": f"{new_material.margin}%"
            }
        }), 201

    except Exception as e:
        print(f"Error in add_material: {e}")  # For debugging
        return jsonify({"error": "Internal Server Error", "details": str(e)}), 500


    
@material_bp.route('/active-material', methods=['GET'])
def get_active_material():
    try:
        # Changed status from "active" to "Released"
        active_material = Material.query.filter_by(status="Released").first()

        if not active_material:
            logging.warning("No released material found.")
            return jsonify({"message": "No released material found."}), 404

        material_data = {
            "material_id": active_material.material_id,
            "title": active_material.title,
            "description": active_material.description,
            "unit_of_measure": active_material.unit_of_measure,
            "current_quantity": str(active_material.current_quantity),
            "minimum_quantity": str(active_material.minimum_quantity),
            "maximum_quantity": str(active_material.maximum_quantity),
            "plant_area_location": active_material.plant_area_location,
            "barcode_id": active_material.barcode_id,
            "status": active_material.status,
            "supplier": active_material.supplier,
            "supplier_contact_info": active_material.supplier_contact_info,
            "notes": active_material.notes,
            "created_at": active_material.created_at,
            "updated_at": active_material.updated_at,
            "margin": f"{active_material.margin}%"  # Add margin to the response
        }

        return jsonify(material_data), 200

    except SQLAlchemyError as e:
        logging.error(f"Database error while fetching released material: {str(e)}")
        return jsonify({"error": "Internal server error. Could not fetch released material."}), 500

    except Exception as e:
        logging.exception(f"Unexpected error occurred: {str(e)}")
        return jsonify({"error": "An unexpected error occurred."}), 500

    
@material_bp.route("/materials/all", methods=["GET"])
def get_all_materials():
    materials = Material.query.all()
    return jsonify(materials_schema.dump(materials)), 200

@material_bp.route('/change-status-to-completed/<int:material_id>', methods=['POST'])
def change_status_to_completed(material_id):
    try:
        # Fetch the material by ID
        material = Material.query.get(material_id)
        if not material:
            logging.warning(f"Material with ID {material_id} not found.")
            return jsonify({"message": f"Material with ID {material_id} not found."}), 404

        # Check if already completed
        if material.status == "completed":
            logging.info(f"Material {material_id} is already marked as completed.")
            return jsonify({"message": f"Material {material_id} is already completed."}), 400

        # Mark the current material as completed
        material.status = "completed"
        db.session.commit()
        logging.info(f"Material {material_id} status updated to 'completed'.")

        # Auto-select and activate the next inactive material (if available)
        next_active_material = Material.query.filter_by(status="inactive").first()
        if next_active_material:
            next_active_material.status = "active"
            db.session.commit()
            logging.info(f"Material {next_active_material.material_id} status updated to 'active'.")
            return jsonify({
                "message": f"Material {material_id} is now completed. "
                           f"Material {next_active_material.material_id} is now active."
            }), 200
        else:
            logging.info(f"Material {material_id} is completed. No inactive material found to activate.")
            return jsonify({
                "message": f"Material {material_id} is now completed, but no inactive material available to activate."
            }), 200

    except SQLAlchemyError as e:
        db.session.rollback()
        logging.error(f"Database error during status change: {str(e)}")
        return jsonify({"error": "Internal server error. Could not update material status."}), 500

    except Exception as e:
        logging.exception(f"Unexpected error occurred: {str(e)}")
        return jsonify({"error": "An unexpected error occurred."}), 500


# ➤ Get Material by Barcode
@material_bp.route("/material/barcode/<string:barcode>", methods=["GET"])
def get_material_by_barcode(barcode):
    try:
        material = Material.query.filter_by(barcode_id=barcode).first()

        if not material:
            return jsonify({"message": "Material not found"}), 404

        return jsonify({
            "material_id": material.material_id,
            "title": material.title,
            "description": material.description,
            "unit_of_measure": material.unit_of_measure,
            "current_quantity": material.current_quantity,
            "minimum_quantity": material.minimum_quantity,
            "maximum_quantity": material.maximum_quantity,
            "plant_area_location": material.plant_area_location,
            "barcode_id": material.barcode_id,
            "status": material.status,
            "supplier": material.supplier,
            "supplier_contact_info": material.supplier_contact_info,
            "notes": material.notes
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@material_bp.route("/materials", methods=["GET"])
def get_materials():
    # Get query parameters from frontend: ?page=1&limit=20
    page = request.args.get("page", 1, type=int)
    limit = request.args.get("limit", 20, type=int)

    # Calculate offset for SQL
    offset = (page - 1) * limit

    # Query only a slice of materials
    materials = Material.query.offset(offset).limit(limit).all()

    # Count total for frontend pagination UI
    total = Material.query.count()

    return jsonify({
        "materials": materials_schema.dump(materials),
        "total": total,
        "page": page,
        "limit": limit
    }), 200


# ➤ Get a specific Material by ID
@material_bp.route("/materials/<int:material_id>", methods=["GET"])
def get_material(material_id):
    material = Material.query.get(material_id)
    if not material:
        return jsonify({"message": "Material not found"}), 404
    return jsonify(material_schema.dump(material)), 200

# ➤ Update a Material
@material_bp.route("/materials/<int:material_id>", methods=["PUT"])
def update_material(material_id):
    material = Material.query.get(material_id)
    if not material:
        return jsonify({"message": "Material not found"}), 404
    
    data = request.get_json()
    material.title = data.get("title", material.title)
    material.description = data.get("description", material.description)
    material.unit_of_measure = data.get("unit_of_measure", material.unit_of_measure)
    material.current_quantity = data.get("current_quantity", material.current_quantity)
    material.minimum_quantity = data.get("minimum_quantity", material.minimum_quantity)
    material.maximum_quantity = data.get("maximum_quantity", material.maximum_quantity)
    material.plant_area_location = data.get("plant_area_location", material.plant_area_location)
    material.barcode_id = data.get("barcode_id", material.barcode_id)
    material.status = data.get("status", material.status)

    db.session.commit()
    return jsonify(material_schema.dump(material)), 200

@material_bp.route("/materials/<int:material_id>", methods=["DELETE"])
def delete_material(material_id):
    try:
        # Fetch the material to ensure it exists
        material = Material.query.get(material_id)
        if not material:
            logger.warning(f"Material with ID {material_id} not found.")
            return jsonify({"error": "Material not found"}), 404

        # Check if the material is referenced in recipe_material table
        referenced_in_recipes = db.session.query(db.func.count()).filter(RecipeMaterial.material_id == material_id).scalar()
        if referenced_in_recipes > 0:
            logger.warning(f"Material with ID {material_id} is referenced in recipe_material table.")
            return jsonify({
                "error": "Cannot delete material because it is referenced in the recipe_material table."
            }), 400

        # Proceed with deletion of transactions and material if no references exist
        try:
            MaterialTransaction.query.filter_by(material_id=material_id).delete()
            db.session.commit()  # Ensure transactions are deleted before deleting material
            logger.info(f"Deleted transactions for Material ID {material_id}")
        except SQLAlchemyError as e:
            db.session.rollback()
            logger.error(f"Error deleting transactions for Material ID {material_id}: {str(e)}")
            return jsonify({"error": "Failed to delete transactions."}), 500

        try:
            db.session.delete(material)
            db.session.commit()
            logger.info(f"Material ID {material_id} and its associated transactions deleted successfully.")
            return jsonify({"message": "Material and associated transactions deleted successfully"}), 200
        except IntegrityError as e:
            db.session.rollback()
            logger.error(f"IntegrityError: Cannot delete Material ID {material_id} due to foreign key constraint: {str(e)}")
            return jsonify({
                "error": "Cannot delete material because it is referenced in another table (foreign key constraint)."
            }), 400
        except SQLAlchemyError as e:
            db.session.rollback()
            logger.error(f"Database error during deletion of Material ID {material_id}: {str(e)}")
            return jsonify({"error": "Database error occurred during deletion."}), 500
    except SQLAlchemyError as e:
        db.session.rollback()
        logger.error(f"Database error during deletion of Material ID {material_id}: {str(e)}")
        return jsonify({"error": "Database error occurred during deletion."}), 500

# ➤ Create a Material Transaction
@material_bp.route("/material-transactions", methods=["POST"])
def create_material_transaction():
    data = request.get_json()
    new_transaction = MaterialTransaction(
        material_id=data["material_id"],
        transaction_type=data["transaction_type"],
        quantity=data["quantity"],
        description=data.get("description")
    )
    db.session.add(new_transaction)
    db.session.commit()
    return jsonify(transaction_schema.dump(new_transaction)), 201

# ➤ Get all Material Transactions
@material_bp.route("/material-transactions", methods=["GET"])
def get_material_transactions():
    transactions = MaterialTransaction.query.all()
    return jsonify(transactions_schema.dump(transactions)), 200

# ➤ Get a specific Material Transaction by ID
@material_bp.route("/material-transactions/<int:transaction_id>", methods=["GET"])
def get_material_transaction(transaction_id):
    transaction = MaterialTransaction.query.get(transaction_id)
    if not transaction:
        return jsonify({"message": "Transaction not found"}), 404
    return jsonify(transaction_schema.dump(transaction)), 200

